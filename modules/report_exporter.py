# ============================================================
#   FaceSecuritySystem — modules/report_exporter.py
#   Module 11: Reports & Export
#   Generates PDF and Excel reports for logs and attendance
# ============================================================

import os
import sys
from datetime import datetime

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
from config.config import EXPORTS_DIR
from database.db_manager import db
from modules.logger import EntryLogger


class ReportExporter:
    """
    Generates PDF and Excel reports.

    Reports available:
        1. Entry Logs Report     — all recognition events
        2. Attendance Report     — daily attendance per user
        3. Attendance Summary    — % per user over date range
        4. Alert Report          — all security alerts
        5. Full Security Report  — combined everything in one PDF

    Usage:
        exporter = ReportExporter()
        path = exporter.export_logs_excel("2024-01-01", "2024-01-31")
        path = exporter.export_logs_pdf("2024-01-01", "2024-01-31")
        path = exporter.export_attendance_pdf("2024-01-15")
        path = exporter.export_summary_excel("2024-01-01", "2024-01-31")
        path = exporter.export_full_pdf("2024-01-01", "2024-01-31")
    """

    def __init__(self):
        os.makedirs(EXPORTS_DIR, exist_ok=True)
        self.logger = EntryLogger()
        print("[EXPORTER] ReportExporter ready.")

    # =========================================================
    #  EXCEL EXPORTS
    # =========================================================

    def export_logs_excel(self, date_from: str = None,
                           date_to: str = None,
                           status: str = None) -> dict:
        """
        Export entry logs to Excel (.xlsx).
        Returns {"success": bool, "path": str, "message": str}
        """
        try:
            import openpyxl
            from openpyxl.styles import (
                Font, PatternFill, Alignment, Border, Side
            )

            logs = self.logger.get_logs(
                date_from=date_from, date_to=date_to, status=status)

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Entry Logs"

            # ── Styles ────────────────────────────────────────
            header_font    = Font(bold=True, color="FFFFFF", size=11)
            header_fill    = PatternFill("solid", fgColor="1F6FEB")
            title_font     = Font(bold=True, size=14, color="1F6FEB")
            center_align   = Alignment(horizontal="center", vertical="center")
            left_align     = Alignment(horizontal="left",   vertical="center")

            thin_border = Border(
                left   = Side(style="thin", color="D0D7DE"),
                right  = Side(style="thin", color="D0D7DE"),
                top    = Side(style="thin", color="D0D7DE"),
                bottom = Side(style="thin", color="D0D7DE"),
            )

            known_fill   = PatternFill("solid", fgColor="1A4226")
            unknown_fill = PatternFill("solid", fgColor="4A1A18")
            spoof_fill   = PatternFill("solid", fgColor="4A3A10")

            # ── Title row ─────────────────────────────────────
            ws.merge_cells("A1:G1")
            ws["A1"] = "Face Security System — Entry Logs Report"
            ws["A1"].font      = title_font
            ws["A1"].alignment = center_align

            ws.merge_cells("A2:G2")
            df_str = date_from or "All"
            dt_str = date_to   or "All"
            ws["A2"] = (f"Period: {df_str}  to  {dt_str}  |  "
                        f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}  |  "
                        f"Total Records: {len(logs)}")
            ws["A2"].font      = Font(size=10, color="8B949E")
            ws["A2"].alignment = center_align

            ws.row_dimensions[1].height = 28
            ws.row_dimensions[2].height = 18

            # ── Header row ────────────────────────────────────
            headers = ["#", "Name", "Date", "Time",
                       "Camera", "Status", "Confidence"]
            widths  = [6, 24, 14, 12, 18, 16, 14]

            for col_i, (h, w) in enumerate(zip(headers, widths), start=1):
                cell = ws.cell(row=3, column=col_i, value=h)
                cell.font      = header_font
                cell.fill      = header_fill
                cell.alignment = center_align
                cell.border    = thin_border
                ws.column_dimensions[
                    openpyxl.utils.get_column_letter(col_i)
                ].width = w

            ws.row_dimensions[3].height = 22

            # ── Data rows ─────────────────────────────────────
            for row_i, log in enumerate(logs, start=4):
                status_val = log.get("status", "")
                conf       = log.get("confidence_score")
                conf_str   = f"{conf:.0%}" if conf else "—"

                row_data = [
                    row_i - 3,
                    log.get("name", ""),
                    str(log.get("entry_date", "")),
                    str(log.get("entry_time", ""))[:8],
                    log.get("camera_label", ""),
                    status_val.title(),
                    conf_str,
                ]

                row_fill = (known_fill   if status_val == "known"
                            else unknown_fill if status_val == "unknown"
                            else spoof_fill)

                for col_i, val in enumerate(row_data, start=1):
                    cell = ws.cell(row=row_i, column=col_i, value=val)
                    cell.alignment = (center_align
                                      if col_i != 2 else left_align)
                    cell.border    = thin_border
                    cell.fill      = row_fill
                    if status_val == "known":
                        cell.font = Font(color="3FB950", size=10)
                    elif status_val == "unknown":
                        cell.font = Font(color="F85149", size=10)
                    else:
                        cell.font = Font(color="D29922", size=10)

                ws.row_dimensions[row_i].height = 18

            # Freeze top 3 rows
            ws.freeze_panes = "A4"

            # Save
            ts   = datetime.now().strftime("%Y%m%d_%H%M%S")
            path = os.path.join(EXPORTS_DIR, f"entry_logs_{ts}.xlsx")
            wb.save(path)

            print(f"[EXPORTER] Excel logs saved: {path}")
            return {"success": True, "path": path,
                    "message": f"Exported {len(logs)} records."}

        except Exception as e:
            print(f"[EXPORTER] Excel error: {e}")
            return {"success": False, "path": None,
                    "message": str(e)}

    def export_attendance_excel(self, date_from: str,
                                 date_to: str) -> dict:
        """Export attendance summary to Excel."""
        try:
            import openpyxl
            from openpyxl.styles import (
                Font, PatternFill, Alignment, Border, Side
            )

            summary = self.logger.get_attendance_summary(
                date_from, date_to)
            users   = summary.get("users", [])

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Attendance Summary"

            header_font  = Font(bold=True, color="FFFFFF", size=11)
            header_fill  = PatternFill("solid", fgColor="1F6FEB")
            title_font   = Font(bold=True, size=14, color="1F6FEB")
            center_align = Alignment(horizontal="center", vertical="center")
            left_align   = Alignment(horizontal="left",   vertical="center")
            thin_border  = Border(
                left   = Side(style="thin", color="D0D7DE"),
                right  = Side(style="thin", color="D0D7DE"),
                top    = Side(style="thin", color="D0D7DE"),
                bottom = Side(style="thin", color="D0D7DE"),
            )

            # Title
            ws.merge_cells("A1:H1")
            ws["A1"] = "Face Security System — Attendance Summary"
            ws["A1"].font      = title_font
            ws["A1"].alignment = center_align
            ws.row_dimensions[1].height = 28

            ws.merge_cells("A2:H2")
            ws["A2"] = (f"Period: {date_from}  to  {date_to}  |  "
                        f"Total Days: {summary['total_days']}  |  "
                        f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}")
            ws["A2"].font      = Font(size=10, color="8B949E")
            ws["A2"].alignment = center_align
            ws.row_dimensions[2].height = 18

            # Header
            headers = ["#", "Name", "Department",
                       "Present", "Late", "Absent",
                       "Total Days", "Attendance %"]
            widths  = [6, 24, 18, 10, 10, 10, 12, 14]

            for col_i, (h, w) in enumerate(zip(headers, widths), start=1):
                cell = ws.cell(row=3, column=col_i, value=h)
                cell.font      = header_font
                cell.fill      = header_fill
                cell.alignment = center_align
                cell.border    = thin_border
                ws.column_dimensions[
                    openpyxl.utils.get_column_letter(col_i)
                ].width = w

            ws.row_dimensions[3].height = 22

            # Data
            for row_i, u in enumerate(users, start=4):
                pct      = u["percentage"]
                pct_fill = PatternFill("solid",
                    fgColor=("1A4226" if pct >= 80
                              else "4A3A10" if pct >= 60
                              else "4A1A18"))
                pct_color = ("3FB950" if pct >= 80
                             else "D29922" if pct >= 60
                             else "F85149")

                row_data = [
                    row_i - 3,
                    u["name"],
                    u["department"],
                    u["present"],
                    u["late"],
                    u["absent"],
                    u["total_days"],
                    f"{pct}%",
                ]

                for col_i, val in enumerate(row_data, start=1):
                    cell = ws.cell(row=row_i, column=col_i, value=val)
                    cell.border    = thin_border
                    cell.alignment = (left_align if col_i == 2
                                      else center_align)
                    if col_i == 8:
                        cell.fill = pct_fill
                        cell.font = Font(
                            color=pct_color, bold=True, size=10)
                    else:
                        cell.font = Font(color="C9D1D9", size=10)

                ws.row_dimensions[row_i].height = 18

            ws.freeze_panes = "A4"

            ts   = datetime.now().strftime("%Y%m%d_%H%M%S")
            path = os.path.join(
                EXPORTS_DIR, f"attendance_{ts}.xlsx")
            wb.save(path)

            print(f"[EXPORTER] Excel attendance saved: {path}")
            return {"success": True, "path": path,
                    "message": f"Exported {len(users)} users."}

        except Exception as e:
            return {"success": False, "path": None,
                    "message": str(e)}

    # =========================================================
    #  PDF EXPORTS
    # =========================================================

    def export_logs_pdf(self, date_from: str = None,
                         date_to: str = None,
                         status: str = None) -> dict:
        """Export entry logs to PDF."""
        try:
            from reportlab.lib.pagesizes  import A4, landscape
            from reportlab.lib            import colors
            from reportlab.lib.styles     import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib.units      import cm
            from reportlab.platypus       import (
                SimpleDocTemplate, Table, TableStyle,
                Paragraph, Spacer
            )

            logs = self.logger.get_logs(
                date_from=date_from, date_to=date_to, status=status)

            ts   = datetime.now().strftime("%Y%m%d_%H%M%S")
            path = os.path.join(EXPORTS_DIR, f"entry_logs_{ts}.pdf")

            doc    = SimpleDocTemplate(
                path,
                pagesize  = landscape(A4),
                leftMargin= 1.5*cm, rightMargin= 1.5*cm,
                topMargin = 1.5*cm, bottomMargin= 1.5*cm
            )
            styles = getSampleStyleSheet()
            story  = []

            # Title
            title_style = ParagraphStyle(
                "title",
                parent    = styles["Heading1"],
                fontSize  = 18,
                textColor = colors.HexColor("#1F6FEB"),
                spaceAfter= 4
            )
            sub_style = ParagraphStyle(
                "sub",
                parent    = styles["Normal"],
                fontSize  = 9,
                textColor = colors.HexColor("#8B949E"),
                spaceAfter= 16
            )

            story.append(Paragraph(
                "Face Security System — Entry Logs Report",
                title_style))
            story.append(Paragraph(
                f"Period: {date_from or 'All'}  to  {date_to or 'All'}  |  "
                f"Total Records: {len(logs)}  |  "
                f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}",
                sub_style))

            # Table data
            col_headers = ["#", "Name", "Date", "Time",
                           "Camera", "Status", "Confidence"]
            table_data  = [col_headers]

            for i, log in enumerate(logs, start=1):
                conf = log.get("confidence_score")
                table_data.append([
                    str(i),
                    str(log.get("name", "")),
                    str(log.get("entry_date", "")),
                    str(log.get("entry_time", ""))[:8],
                    str(log.get("camera_label", "")),
                    str(log.get("status", "")).title(),
                    f"{conf:.0%}" if conf else "—",
                ])

            col_widths = [1*cm, 5*cm, 3*cm, 2.5*cm,
                          4*cm, 3*cm, 2.5*cm]

            table = Table(table_data, colWidths=col_widths,
                          repeatRows=1)

            table.setStyle(TableStyle([
                # Header
                ("BACKGROUND",  (0, 0), (-1, 0),
                 colors.HexColor("#1F6FEB")),
                ("TEXTCOLOR",   (0, 0), (-1, 0), colors.white),
                ("FONTNAME",    (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE",    (0, 0), (-1, 0), 10),
                ("ALIGN",       (0, 0), (-1, 0), "CENTER"),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1),
                 [colors.HexColor("#161B22"),
                  colors.HexColor("#1A1F27")]),
                ("TEXTCOLOR",   (0, 1), (-1, -1),
                 colors.HexColor("#C9D1D9")),
                ("FONTSIZE",    (0, 1), (-1, -1), 9),
                ("ALIGN",       (0, 1), (-1, -1), "CENTER"),
                ("ALIGN",       (1, 1), (1, -1),  "LEFT"),
                ("GRID",        (0, 0), (-1, -1),
                 0.5, colors.HexColor("#30363D")),
                ("ROWHEIGHT",   (0, 0), (-1, -1), 20),
                ("TOPPADDING",  (0, 0), (-1, -1), 5),
                ("BOTTOMPADDING",(0,0), (-1, -1), 5),
            ]))

            story.append(table)
            doc.build(story)

            print(f"[EXPORTER] PDF logs saved: {path}")
            return {"success": True, "path": path,
                    "message": f"Exported {len(logs)} records to PDF."}

        except Exception as e:
            print(f"[EXPORTER] PDF error: {e}")
            return {"success": False, "path": None,
                    "message": str(e)}

    def export_attendance_pdf(self, date_from: str,
                               date_to: str) -> dict:
        """Export attendance summary to PDF."""
        try:
            from reportlab.lib.pagesizes import A4
            from reportlab.lib           import colors
            from reportlab.lib.styles    import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib.units     import cm
            from reportlab.platypus      import (
                SimpleDocTemplate, Table, TableStyle,
                Paragraph, Spacer
            )

            summary = self.logger.get_attendance_summary(
                date_from, date_to)
            users   = summary.get("users", [])

            ts   = datetime.now().strftime("%Y%m%d_%H%M%S")
            path = os.path.join(
                EXPORTS_DIR, f"attendance_{ts}.pdf")

            doc    = SimpleDocTemplate(
                path, pagesize=A4,
                leftMargin=1.5*cm, rightMargin=1.5*cm,
                topMargin=1.5*cm,  bottomMargin=1.5*cm
            )
            styles = getSampleStyleSheet()
            story  = []

            title_style = ParagraphStyle(
                "t", parent=styles["Heading1"],
                fontSize=18,
                textColor=colors.HexColor("#1F6FEB"),
                spaceAfter=4)
            sub_style = ParagraphStyle(
                "s", parent=styles["Normal"],
                fontSize=9,
                textColor=colors.HexColor("#8B949E"),
                spaceAfter=16)

            story.append(Paragraph(
                "Face Security System — Attendance Report",
                title_style))
            story.append(Paragraph(
                f"Period: {date_from}  to  {date_to}  |  "
                f"Total Days: {summary['total_days']}  |  "
                f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}",
                sub_style))

            col_headers = ["#", "Name", "Department",
                           "Present", "Late", "Absent",
                           "Days", "Attendance %"]
            table_data  = [col_headers]

            for i, u in enumerate(users, start=1):
                table_data.append([
                    str(i),
                    u["name"],
                    u["department"],
                    str(u["present"]),
                    str(u["late"]),
                    str(u["absent"]),
                    str(u["total_days"]),
                    f"{u['percentage']}%",
                ])

            col_widths = [1*cm, 5*cm, 4*cm, 2*cm,
                          2*cm, 2*cm, 2*cm, 3*cm]

            table = Table(table_data, colWidths=col_widths,
                          repeatRows=1)

            table.setStyle(TableStyle([
                ("BACKGROUND",  (0,0), (-1,0),
                 colors.HexColor("#1F6FEB")),
                ("TEXTCOLOR",   (0,0), (-1,0), colors.white),
                ("FONTNAME",    (0,0), (-1,0), "Helvetica-Bold"),
                ("FONTSIZE",    (0,0), (-1,0), 10),
                ("ALIGN",       (0,0), (-1,0), "CENTER"),
                ("ROWBACKGROUNDS", (0,1), (-1,-1),
                 [colors.HexColor("#161B22"),
                  colors.HexColor("#1A1F27")]),
                ("TEXTCOLOR",   (0,1), (-1,-1),
                 colors.HexColor("#C9D1D9")),
                ("FONTSIZE",    (0,1), (-1,-1), 9),
                ("ALIGN",       (0,1), (-1,-1), "CENTER"),
                ("ALIGN",       (1,1), (2,-1),  "LEFT"),
                ("GRID",        (0,0), (-1,-1),
                 0.5, colors.HexColor("#30363D")),
                ("ROWHEIGHT",   (0,0), (-1,-1), 22),
                ("TOPPADDING",  (0,0), (-1,-1), 5),
                ("BOTTOMPADDING",(0,0),(-1,-1), 5),
            ]))

            story.append(table)
            doc.build(story)

            print(f"[EXPORTER] PDF attendance saved: {path}")
            return {"success": True, "path": path,
                    "message": f"Exported {len(users)} users to PDF."}

        except Exception as e:
            return {"success": False, "path": None,
                    "message": str(e)}

    def export_alerts_excel(self, date_from: str = None,
                             date_to: str = None) -> dict:
        """Export alerts to Excel."""
        try:
            import openpyxl
            from openpyxl.styles import (
                Font, PatternFill, Alignment, Border, Side
            )

            query  = "SELECT * FROM alerts WHERE 1=1"
            params = []
            if date_from:
                query += " AND alert_date >= %s"
                params.append(date_from)
            if date_to:
                query += " AND alert_date <= %s"
                params.append(date_to)
            query += " ORDER BY created_at DESC"
            alerts = db.fetch_all(query, tuple(params))

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Alerts"

            header_font  = Font(bold=True, color="FFFFFF", size=11)
            header_fill  = PatternFill("solid", fgColor="F85149")
            title_font   = Font(bold=True, size=14, color="F85149")
            center_align = Alignment(
                horizontal="center", vertical="center")
            thin_border  = Border(
                left   = Side(style="thin", color="D0D7DE"),
                right  = Side(style="thin", color="D0D7DE"),
                top    = Side(style="thin", color="D0D7DE"),
                bottom = Side(style="thin", color="D0D7DE"))

            ws.merge_cells("A1:F1")
            ws["A1"] = "Face Security System — Alerts Report"
            ws["A1"].font      = title_font
            ws["A1"].alignment = center_align
            ws.row_dimensions[1].height = 28

            ws.merge_cells("A2:F2")
            ws["A2"] = (f"Total: {len(alerts)}  |  "
                        f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}")
            ws["A2"].font      = Font(size=10, color="8B949E")
            ws["A2"].alignment = center_align
            ws.row_dimensions[2].height = 18

            headers = ["#", "Alert Type", "Date", "Time",
                       "Camera", "Reviewed"]
            widths  = [6, 24, 14, 12, 18, 12]

            for col_i, (h, w) in enumerate(
                    zip(headers, widths), start=1):
                cell = ws.cell(row=3, column=col_i, value=h)
                cell.font      = header_font
                cell.fill      = header_fill
                cell.alignment = center_align
                cell.border    = thin_border
                ws.column_dimensions[
                    openpyxl.utils.get_column_letter(col_i)
                ].width = w

            for row_i, alert in enumerate(alerts, start=4):
                reviewed  = alert.get("is_reviewed", 0)
                row_fill  = PatternFill(
                    "solid",
                    fgColor="1A4226" if reviewed else "4A1A18")
                row_data  = [
                    row_i - 3,
                    str(alert.get("alert_type","")).replace(
                        "_"," ").title(),
                    str(alert.get("alert_date","")),
                    str(alert.get("alert_time",""))[:8],
                    str(alert.get("camera_label","")),
                    "Yes" if reviewed else "No",
                ]
                for col_i, val in enumerate(row_data, start=1):
                    cell = ws.cell(row=row_i, column=col_i,
                                    value=val)
                    cell.border    = thin_border
                    cell.alignment = center_align
                    cell.fill      = row_fill
                    cell.font      = Font(
                        color="3FB950" if reviewed else "F85149",
                        size=10)
                ws.row_dimensions[row_i].height = 18

            ws.freeze_panes = "A4"

            ts   = datetime.now().strftime("%Y%m%d_%H%M%S")
            path = os.path.join(EXPORTS_DIR, f"alerts_{ts}.xlsx")
            wb.save(path)

            return {"success": True, "path": path,
                    "message": f"Exported {len(alerts)} alerts."}

        except Exception as e:
            return {"success": False, "path": None,
                    "message": str(e)}

    # =========================================================
    #  LIST EXPORTS
    # =========================================================

    def get_export_files(self) -> list:
        """Return list of all exported files with metadata."""
        files = []
        if not os.path.exists(EXPORTS_DIR):
            return files
        for fname in sorted(os.listdir(EXPORTS_DIR), reverse=True):
            fpath = os.path.join(EXPORTS_DIR, fname)
            if os.path.isfile(fpath):
                size = os.path.getsize(fpath)
                files.append({
                    "name"    : fname,
                    "path"    : fpath,
                    "size"    : self._format_size(size),
                    "modified": datetime.fromtimestamp(
                        os.path.getmtime(fpath)
                    ).strftime("%Y-%m-%d %H:%M"),
                    "type"    : "Excel" if fname.endswith(".xlsx")
                                else "PDF" if fname.endswith(".pdf")
                                else "Other"
                })
        return files

    @staticmethod
    def _format_size(size_bytes: int) -> str:
        if size_bytes < 1024:
            return f"{size_bytes} B"
        elif size_bytes < 1024 * 1024:
            return f"{size_bytes / 1024:.1f} KB"
        else:
            return f"{size_bytes / (1024*1024):.1f} MB"


# Module-level singleton
report_exporter = ReportExporter()